from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import os

def generate_excel_report(data):
    # Gera um relatório em Excel com base nos dados fornecidos
    report_filename = f"report_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    report_path = os.path.join(os.path.dirname(__file__), report_filename)
    
    wb = Workbook()
    ws = wb.active
    
    # Adiciona os cabeçalhos
    headers = list(data[0].keys()) + ["Quantidade de atendimentos do dia"]
    ws.append(headers)
    
    # Contador de atendimentos por data
    date_count = {}
    
    for row in data:
        date = row["Data"]
        if date in date_count:
            date_count[date] += 1
        else:
            date_count[date] = 1
    
    # Insere os dados nas colunas correspondentes
    row_index = 2  # Começa na segunda linha após os cabeçalhos
    for row in data:
        values = list(row.values())
        date = row["Data"]
        values.append(date_count[date])
        ws.append(values)
        row_index += 1
    
    # Mescla as células para "Quantidade de atendimentos do dia"
    current_row = 2  # Começa na segunda linha após os cabeçalhos
    while current_row <= ws.max_row:
        current_date = ws[f"H{current_row}"].value
        span = 0
        
        # Conta quantas linhas têm a mesma data
        for r in range(current_row, ws.max_row + 1):
            if ws[f"H{r}"].value == current_date:
                span += 1
            else:
                break
        
        if span > 1:
            cell_to_merge = f"J{current_row}:J{current_row + span - 1}"
            ws.merge_cells(cell_to_merge)
            ws[f"J{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        current_row += span
    
    # Formata as células para centralizar o texto vertical e horizontalmente no cabeçalho
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    wb.save(report_path)
    
    return report_path
